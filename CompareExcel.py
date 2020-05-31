from flask import Flask
from flask_restful import Resource,Api
import openpyxl,xlrd
import os

app = Flask(__name__)
api = Api(app)

class ComExcel(Resource):

    # 类列表
    class_list = list()

    def get(self, one_excel, two_excel):
        dir_name = os.path.realpath(r'C:\Users\Acer\Desktop')
        one_path = os.path.join(dir_name, one_excel)
        two_path = os.path.join(dir_name, two_excel)
        # one_wb = openpyxl.load_workbook(one_excel)
        # two_wb = openpyxl.load_workbook(two_excel)
        # sheet_cell = list()
        # for name in one_wb.sheetnames:
        #     wb = one_wb[name]
        try:
            wb_one = xlrd.open_workbook(one_path)
            wb_tow = xlrd.open_workbook(two_path)
        except FileNotFoundError as e:
            return 200,{
                "msg":"该目录不存在"
            }
        # 比较两个excel的高度和宽度
        one_sheet = wb_one.sheet_by_index(0)
        two_sheet = wb_tow.sheet_by_index(0)
        one_rows = one_sheet.nrows
        two_rows = two_sheet.nrows
        one_cols = one_sheet.ncols
        two_cols = two_sheet.ncols
        if one_rows !=two_rows and one_cols != two_cols:
            return {
                "msg":"两个excel表宽度或者高度不一致"
            }
        # 比较两个excel的内容
        for one_row in range(one_rows):
            one_line_values = one_sheet.row_values(one_row)
            for two_line in range(one_row, one_row+1):
                two_line_values = two_sheet.row_values(two_line)
                self.compare(one_line_values, two_line_values, one_row+1)

        if len(ComExcel.class_list) == 0:
            return {
                "msg":"两个表完全相同"
            }
        return ComExcel.class_list


    def compare(self, list1, list2, line):
        if len(list1) == len(list2):
            print("第{}行长度相同".format(line))
            for i in range(len(list2)):
                if list1[i] == list2[i]:
                    # print("第{}行数据相同".format(line))
                    continue
                else:
                    self.class_list.append({
                        "result":"第{}行数据不同".format(line),
                        "msg":{
                            "dif_line1":list1[i],
                            "dif_line2":list2[i]
                        }
                    })
            # return {
            #     "msg":"第{}行数据不同".format(line)
            # }

api.add_resource(ComExcel, "/<one_excel>/<two_excel>")

if __name__ == '__main__':
    app.run(debug=True)